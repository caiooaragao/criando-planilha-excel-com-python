
# template de criaçao de planilha no excel com python

import openpyxl

# dados teste
dados = [1, 'vaca', 'cavalo', 3, 'vaca']

# cria do uma planilha no excel e populando ela com python

book = openpyxl.Workbook()

# criando a planilha
book.create_sheet('planilhaJu')


# abrindo o arquivo da planilha
book = openpyxl.load_workbook('planilha de ju.xlsx')

# selecionando a página a ser usada dentro da planilha
dados_empresas = book['planilhaJu']


# dados a serem inseridos na ultima linha
dados_empresas.append(
    [dados[0], dados[1], dados[2], dados[3], dados[4]])


# salvando os dados que foram inseridos na planilha
book.save('planilha de ju.xlsx')
